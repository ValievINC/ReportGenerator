import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import asyncio
from CreateDictionary import create_dictionary, get_projects, get_date_statistic, get_month_statistic

# Styles
text_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
text_style_manuscript = Alignment(horizontal='center', vertical='top', wrap_text=True)
text_style_human = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
text_style_groups = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

font1 = Font(name='Verdana', size=6, bold=False, italic=False, color='000000')
font2 = Font(name='Verdana', size=8, bold=False, italic=False, color='000000')
font3 = Font(name='Verdana', size=12, bold=True, italic=False, color='000000')
font4 = Font(name='Verdana', size=7, bold=True, italic=False, color='000000')

red_fill = PatternFill(start_color='A52A2A', end_color='A52A2A', fill_type='solid')
green_fill = PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid')
blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')

border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Date Dictionary
months = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь',
}


def create_employee_sheet(workbook, name, month):
    sheet = workbook.create_sheet()
    sheet.title = name

    # Columns Width
    sheet.column_dimensions['A'].width = 16
    for column_index in range(2, len(month) + 1):
        column_letter = get_column_letter(column_index)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = 3

    # Rows Height
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 14
    sheet.row_dimensions[3].height = 44

    # Style
    cell_range = sheet[f'A1:{get_column_letter(len(month) + 1)}3']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    cell_range = sheet[f'A3:{get_column_letter(len(month) + 1)}3']
    for row in cell_range:
        for cell in row:
            cell.font = font2

    # Data
    sheet['A1'] = 'Фамилия, инициалы'
    sheet['A3'] = name

    for column_index in range(2, len(month) + 2):
        column_letter = get_column_letter(column_index)
        cell = sheet[f'{column_letter}{1}']
        cell.value = int(column_index - 1)
        cell = sheet[f'{column_letter}{2}']
        cell.value = month[column_index-1]

    # Merging
    sheet.merge_cells('A1:A2')


def fill_page(sheet, dictionary, name, date_stat):
    letter = get_column_letter(date_stat['day'] + 1)
    if 'ЭП-600 НКНХ' in dictionary[name]['group_and_hours'].keys():
        sheet[f'{letter}3'] = float(dictionary[name]['group_and_hours']['ЭП-600 НКНХ'])


def fill_employees_page(workbook, dictionary, date_stat, month):
    for name in dictionary.keys():
        if name not in workbook.sheetnames:
            create_employee_sheet(workbook, name, month)
            fill_page(sheet=workbook[name], dictionary=dictionary, name=name, date_stat=date_stat)
        else:
            sheet = workbook[name]
            fill_page(sheet=sheet, dictionary=dictionary, name=name, date_stat=date_stat)


def collect_information(workbook):
    employees_info = {}
    for page in workbook.worksheets[1:]:
        name = page.title
        employees_info[name] = {}
        for cell_column in range(2, page.max_column + 1):
            letter = get_column_letter(cell_column)
            employees_info[name][page[f'{letter}1'].value] = page[f'{letter}3'].value
    return employees_info


def create_summary_page(workbook, month):
    sheet = workbook.create_sheet()
    sheet.title = 'Сводная таблица'
    workbook.move_sheet(sheet, offset=-(len(workbook.sheetnames) - 1))

    # Columns Width
    sheet.column_dimensions['A'].width = 16
    for column_index in range(2, len(month) + 2):
        column_letter = get_column_letter(column_index)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = 3

    # Rows Height
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 14

    # Data
    sheet['A1'] = 'Фамилия, инициалы'
    for column_index in range(2, len(month) + 2):
        column_letter = get_column_letter(column_index)
        cell = sheet[f'{column_letter}{1}']
        cell.value = int(column_index - 1)
        cell = sheet[f'{column_letter}{2}']
        cell.value = month[column_index - 1]

    employees_data = collect_information(workbook)
    print(employees_data)

    for name in employees_data.keys():
        current_row = sheet.max_row + 1
        sheet[f'A{current_row}'] = name
        for day in employees_data[name].keys():
            letter = get_column_letter(day + 1)
            sheet[f'{letter}{current_row}'] = employees_data[name][day]

    # Style
    cell_range = sheet[f'A1:{get_column_letter(len(month) + 1)}{sheet.max_row}']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    cell_range = sheet[f'A1:{get_column_letter(len(month) + 1)}2']
    for row in cell_range:
        for cell in row:
            cell.border = Border(
                                left=Side(border_style='thin', color='000000'),
                                right=Side(border_style='thin', color='000000'),
                                top=Side(border_style='thin', color='000000'),
                                bottom=Side(border_style='double', color='000000')
                            )

    # Merging
    sheet.merge_cells('A1:A2')

    sheet.freeze_panes = 'AG3'


def find_cell(sheet, value):
    found_cell = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                found_cell = cell
    return found_cell


def sort_pages(workbook):
    sorted_sheets = sorted(workbook._sheets, key=lambda sheet: sheet.title.split()[-1])
    workbook._sheets = sorted_sheets


def create_report(report, selected_date, staff):
    workbook = load_workbook(report)

    date_stat = get_date_statistic(selected_date)
    month = get_month_statistic(selected_date)

    if 'Лист1' in workbook.sheetnames:
        del workbook['Лист1']
    #
    if 'Сводная таблица' in workbook.sheetnames:
        del workbook['Сводная таблица']
    #
    # if 'Проекты' in workbook.sheetnames:
    #     del workbook['Проекты']

    df = pd.read_excel(staff, usecols=['ФИО', ])
    df[['Фамилия', 'Имя', 'Отчество']] = df['ФИО'].str.split(' ', expand=True)
    df['ИФ'] = df['Имя'] + ' ' + df['Фамилия']
    df = df.dropna()
    employees = df['ИФ'].to_numpy()

    for employee in employees:
        if employee not in workbook.sheetnames:
            create_employee_sheet(workbook, employee, month)

    dictionary = asyncio.run(create_dictionary(selected_date, employees))
    fill_employees_page(workbook, dictionary, date_stat, month)

    sort_pages(workbook)

    create_summary_page(workbook, month)

    workbook.save(report)
