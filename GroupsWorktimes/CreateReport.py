import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import asyncio
from CreateDictionary import create_dictionary, get_projects

# Styles
text_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
text_style_manuscript = Alignment(horizontal='center', vertical='top', wrap_text=True)
text_style_human = Alignment(horizontal='left', vertical='bottom', wrap_text=True)
text_style_groups = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

font1 = Font(name='Verdana', size=6, bold=False, italic=False, color='000000')
font2 = Font(name='Verdana', size=8, bold=False, italic=False, color='000000')
font3 = Font(name='Verdana', size=12, bold=True, italic=False, color='000000')
font4 = Font(name='Verdana', size=7, bold=True, italic=False, color='000000')

red_fill = PatternFill(start_color='A52A2A', end_color='A52A2A', fill_type='solid')
green_fill = PatternFill(start_color='32CD32', end_color='32CD32', fill_type='solid')
blue_fill = PatternFill(start_color='87CEFA', end_color='87CEFA', fill_type='solid')

border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)

# Date Dictionary
months = {
    1: 'Январь',
    2: 'Февраль',
    3: 'Март',
    4: 'Апрель',
    5: 'Май',
    6: 'Июнь',
    7: 'Июль',
    8: 'Август',
    9: 'Сентябрь',
    10: 'Октябрь',
    11: 'Ноябрь',
    12: 'Декабрь',
}


def create_employee_sheet(workbook, employee):
    sheet = workbook.create_sheet()
    sheet.title = employee

    # Column Width
    sheet.column_dimensions['A'].width = 36
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 16
    sheet.column_dimensions['D'].width = 16
    sheet.column_dimensions['E'].width = 16

    # Rows Height
    sheet.row_dimensions[1].height = 44
    sheet.row_dimensions[2].height = 44

    # Style
    cell_range = sheet['A1:E2']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font4
            cell.border = border

    # Data
    sheet['A1'] = employee
    sheet['A2'] = "Группа"
    sheet['B2'] = 'Время затрачено'
    sheet['C2'] = '% От общего затраченного времени'
    sheet['D2'] = 'Доля от общего затраченного времени'
    sheet['E2'] = 'Доля от плана'
    sheet['B3'] = 0

    # Merging
    sheet.merge_cells('A1:E1')


def fill_page(sheet, dictionary, name):
    sheet['M1'] = dictionary[name]['city']
    sheet['M2'] = dictionary[name]['work_position']
    sheet['M1'].font = Font(color='FFFFFFFF')
    sheet['M2'].font = Font(color='FFFFFFFF')

    cell_value = sheet[f'B{sheet.max_row}'].value
    if cell_value is not None and isinstance(cell_value, (int, float, str)):
        total_hours = float(cell_value)
    else:
        total_hours = 0.0

    sheet.delete_rows(sheet.max_row, amount=1)

    max_row = sheet.max_row
    current_groups = []

    total_hours += float(dictionary[name]['total_time'])

    for row in range(3, max_row + 1):
        current_groups.append(sheet[f'A{row}'].value)

    for group in dictionary[name]['group_and_hours']:
        if group not in current_groups:
            sheet[f'A{max_row + 1}'] = group
            sheet[f'B{max_row + 1}'] = dictionary[name]['group_and_hours'][group]
            max_row = sheet.max_row
        else:
            index = current_groups.index(group)
            pos = index + 3
            sheet[f'B{pos}'] = float(sheet[f'B{pos}'].value) + dictionary[name]['group_and_hours'][group]

    for row in range(3, max_row + 1):
        sheet[f'C{row}'] = round((float(sheet[f'B{row}'].value) / total_hours) * 100, 1)

    for row in range(3, max_row + 1):
        sheet[f'D{row}'] = round(float(sheet[f'B{row}'].value) / total_hours, 3)

    sheet[f'B{max_row + 1}'] = total_hours


def fill_employees_page(workbook, dictionary):
    for name in dictionary.keys():
        if name not in workbook.sheetnames:
            create_employee_sheet(workbook, name)
            fill_page(sheet=workbook[name], dictionary=dictionary, name=name)
        else:
            sheet = workbook[name]
            fill_page(sheet=sheet, dictionary=dictionary, name=name)


def collect_groups(workbook):
    group_names = []
    for page in workbook.worksheets[1:]:
        max_row = page.max_row
        for row in range(3, max_row):
            group_names.append(page[f'A{row}'].value)
    group_names = set(group_names)
    group_names = sorted(list(group_names))
    result = []
    for name in reversed(group_names):
        if name in ['Разработка P&ID БХК в Diagrams', 'Автоматизация проектирования', 'Административные работы', 'Курсы обучения']:
            result.append(name)
        else:
            result.insert(0, name)
    return result


def collect_information(workbook):
    employees_info = {}
    for page in workbook.worksheets[1:]:
        name = page.title
        employees_info[name] = {'work_position': '', 'group_and_hours': {}, 'city': ''}
        max_row = page.max_row
        for row in range(3, max_row):
            employees_info[name]['group_and_hours'][page[f'A{row}'].value] = float(page[f'C{row}'].value) / 100
        employees_info[name]['work_position'] = page['M2'].value
        employees_info[name]['city'] = page['M1'].value
    print(employees_info)
    result = {}
    for name in employees_info:
        city = employees_info[name]['city']
        if city not in result:
            result[city] = {}
        result[city][name] = {
            'group_and_hours': employees_info[name]['group_and_hours'],
            'work_position': employees_info[name]['work_position']
        }

    return result


def create_resources_page(workbook, selected_date):
    sheet = workbook.create_sheet()
    sheet.title = '1-Ресурсы'
    workbook.move_sheet(sheet, offset=-(len(workbook.sheetnames)-1))

    # Columns Width
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 36
    sheet.column_dimensions['C'].width = 36
    sheet.column_dimensions['D'].width = 24

    # Rows Heights
    sheet.row_dimensions[1].height = 15
    sheet.row_dimensions[2].height = 30

    # Data
    month = months[selected_date.month]
    year = selected_date.year
    sheet['E1'] = f'Доля участия в проекте, % рабочего времени в месяц {month} {year}'
    sheet['A2'] = '№'
    sheet['B2'] = 'ФИО'
    sheet['C2'] = 'Дисциплина'
    sheet['D2'] = 'Должность'

    group_names = collect_groups(workbook)

    for i in range(len(group_names)):
        letter = get_column_letter(i + 5)
        sheet.column_dimensions[f'{letter}'].width = 18
        sheet[f'{letter}2'] = group_names[i]
    sheet[f'{get_column_letter(len(group_names) + 5)}2'] = 'Итог'

    information = collect_information(workbook)

    index = 1
    for city in information.keys():
        max_row = sheet.max_row
        sheet[f'B{max_row + 1}'] = city
        sheet[f'B{max_row + 1}'].font = font3
        sheet[f'B{max_row + 1}'].fill = green_fill
        sheet[f'C{max_row + 1}'].fill = green_fill
        sheet[f'D{max_row + 1}'].fill = green_fill
        for name in information[city]:
            max_row = sheet.max_row
            sheet[f'A{max_row + 1}'] = index
            sheet[f'B{max_row + 1}'] = name
            sheet[f'D{max_row + 1}'] = information[city][name]['work_position']
            index += 1
            result = 0
            for i in range(len(group_names)):
                letter = get_column_letter(i + 5)
                sheet[f'{letter}{max_row + 1}'] = information[city][name]['group_and_hours'].get(sheet[f'{letter}2'].value, 0)
                if sheet[f'{letter}{max_row + 1}'].value > 0:
                    sheet[f'{letter}{max_row + 1}'].fill = blue_fill
                result += float(sheet[f'{letter}{max_row + 1}'].value)
            sheet[f'{get_column_letter(len(group_names) + 5)}{max_row + 1}'] = round(result)

    # Style
    max_row = sheet.max_row
    max_column = sheet.max_column
    cell_range = sheet[f'A1:{get_column_letter(max_column)}{max_row}']
    for row in cell_range:
        for cell in row:
            cell.border = border

    sheet['B2'].alignment = text_style
    sheet['C2'].alignment = text_style
    sheet['D2'].alignment = text_style

    cell_range = sheet[f'A1:{get_column_letter(max_column)}2']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style_groups

    # Merging
    sheet.merge_cells(f'E1:{get_column_letter(max_column)}1')

    sheet.freeze_panes = 'C3'


def collect_parts(workbook):
    resources_sheet = workbook['1-Ресурсы']
    parts_info = {}
    max_row = resources_sheet.max_row
    max_column = resources_sheet.max_column

    for column in range(5, max_column):
        total_part = 0
        for row in range(3, max_row + 1):
            cell_value = resources_sheet[f'{get_column_letter(column)}{row}'].value
            if cell_value is not None:
                total_part += float(cell_value)
        parts_info[resources_sheet[f'{get_column_letter(column)}2'].value] = total_part

    print('PARTS INFO: ', parts_info)
    return parts_info


def create_projects_page(workbook):
    sheet = workbook.create_sheet()
    sheet.title = 'Проекты'
    workbook.move_sheet(sheet, offset=-(len(workbook.sheetnames)-1))

    # Column Width
    sheet.column_dimensions['A'].width = 12
    sheet.column_dimensions['B'].width = 12
    sheet.column_dimensions['C'].width = 12
    sheet.column_dimensions['D'].width = 24
    sheet.column_dimensions['E'].width = 8
    sheet.column_dimensions['F'].width = 12
    sheet.column_dimensions['G'].width = 12
    sheet.column_dimensions['H'].width = 6

    # Rows Heights
    sheet.row_dimensions[1].height = 15
    sheet.row_dimensions[2].height = 45

    # Data
    sheet['H1'] = 'по дисциплинам, доли чел/мес.'
    sheet['A2'] = 'Проект'
    sheet['B2'] = 'Код проекта'
    sheet['C2'] = 'Титул'
    sheet['D2'] = 'Общее название'
    sheet['E2'] = 'Пояснение'
    sheet['F2'] = 'Исполнители, суммарная доля чел/мес'
    sheet['G2'] = '% от общего численного состава'
    sheet['H2'] = '3D'

    parts = collect_parts(workbook)
    groups = asyncio.run(get_projects())

    total = 0
    for group in groups.keys():
        max_row = sheet.max_row
        sheet[f'A{max_row + 1}'] = group
        this_row = max_row + 1
        for field in groups[group]:
            sheet[f'C{this_row}'] = field['title'].split(' - ')[0]
            if f'{group} {field["title"].split(" - ")[0]}' in parts.keys():
                sheet[f'H{this_row}'] = parts[f'{group} {field["title"].split(" - ")[0]}']
                total += float(sheet[f'H{this_row}'].value)
            this_row += 1
        if group in parts.keys():
            sheet[f'A{this_row}'] = group
            sheet[f'H{this_row}'] = parts[group]
            total += float(sheet[f'H{this_row}'].value)
        if this_row > max_row + 1:
            sheet.merge_cells(f'A{max_row + 1}:A{this_row - 1}')
            sheet.merge_cells(f'B{max_row + 1}:B{this_row - 1}')

    unuseful = 0
    for group in parts.keys():
        if group in ['Разработка P&ID БХК в Diagrams', 'Курсы обучения', 'Административные работы', 'Автоматизация проектирования', 'Автоматизация проектирования 90100']:
            unuseful += float(parts[group])
            total += float(parts[group])

    max_row = sheet.max_row
    cell_range = sheet[f'H3:H{sheet.max_row - 1}']
    part_to_add = unuseful / (max_row - 2) if max_row != 3 else unuseful
    for row in cell_range:
        for cell in row:
            if cell.value is not None:
                cell.value += part_to_add
            else:
                cell.value = part_to_add

    max_row = sheet.max_row
    sheet[f'A{max_row + 1}'] = 'Итого:'
    sheet[f'H{max_row + 1}'] = round(total)

    # Style
    max_row = sheet.max_row
    max_column = sheet.max_column
    cell_range = sheet[f'A3:{get_column_letter(max_column)}{max_row}']
    for row in cell_range:
        for cell in row:
            cell.border = border
            cell.alignment = text_style

    cell_range = sheet[f'A2:{get_column_letter(max_column)}2']
    for row in cell_range:
        for cell in row:
            cell.border = border
            cell.alignment = text_style
            cell.font = font4

    # Merging:
    sheet.merge_cells(f'A{max_row}:E{max_row}')


def find_cell(sheet, value):
    found_cell = None
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == value:
                found_cell = cell
    return found_cell


def sort_pages(workbook):
    sorted_sheets = sorted(workbook._sheets, key=lambda sheet: sheet.title.split()[-1])
    workbook._sheets = sorted_sheets


def create_report(report, selected_date, staff):
    workbook = load_workbook(report)

    if 'Лист1' in workbook.sheetnames:
        del workbook['Лист1']

    if '1-Ресурсы' in workbook.sheetnames:
        del workbook['1-Ресурсы']

    if 'Проекты' in workbook.sheetnames:
        del workbook['Проекты']

    df = pd.read_excel(staff, usecols=['ФИО', ])
    df[['Фамилия', 'Имя', 'Отчество']] = df['ФИО'].str.split(' ', expand=True)
    df['ИФ'] = df['Имя'] + ' ' + df['Фамилия']
    df = df.dropna()
    employees = df['ИФ'].to_numpy()

    for employee in employees:
        if employee not in workbook.sheetnames:
            create_employee_sheet(workbook, employee)

    dictionary = asyncio.run(create_dictionary(selected_date, employees))
    fill_employees_page(workbook, dictionary)

    sort_pages(workbook)

    create_resources_page(workbook, selected_date)
    create_projects_page(workbook)

    workbook.save(report)
