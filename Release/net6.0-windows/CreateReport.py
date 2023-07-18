# encoding: windows-1251

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter

from CreateDataframe import create_dataframe


# Styles
text_style = Alignment(horizontal='center', vertical='center', wrap_text=True)
text_style_manuscript = Alignment(horizontal='center', vertical='top', wrap_text=True)
text_style_human = Alignment(horizontal='left', vertical='bottom', wrap_text=True)

font1 = Font(name='Verdana', size=6, bold=False, italic=False, color='000000')
font2 = Font(name='Verdana', size=8, bold=False, italic=False, color='000000')
font3 = Font(name='Verdana', size=12, bold=True, italic=False, color='000000')
font4 = Font(name='Verdana', size=7, bold=True, italic=False, color='000000')

red_fill = PatternFill(start_color='FA8072', end_color='A52A2A', fill_type='solid')

border = Border(
    left=Side(border_style='thin', color='000000'),
    right=Side(border_style='thin', color='000000'),
    top=Side(border_style='thin', color='000000'),
    bottom=Side(border_style='thin', color='000000')
)


# Painting
def create_employee_sheet(workbook, data):
    sheet = workbook.create_sheet()
    sheet.title = data['Сотрудник']

    # Columns Width
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 16
    sheet.column_dimensions['C'].width = 32
    for column_index in range(4, 35):
        column_letter = get_column_letter(column_index)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = 3

    # Rows Height
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 28
    sheet.row_dimensions[3].height = 10
    sheet.row_dimensions[4].height = 44

    cell_range = sheet['A1:AH4']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    cell_range = sheet['A4:C4']
    for row in cell_range:
        for cell in row:
            cell.font = font2

    # Data
    sheet['A1'] = 'Номер\n по\n порядку'
    sheet['B1'] = 'Фамилия, инициалы,\n должность\n (специальность,\n профессия)'
    sheet['C1'] = 'Табельный\n номер'
    sheet['D1'] = 'Отметки о явках и неявках на работу по числам месяца'
    sheet['A3'] = 1
    sheet['B3'] = 2
    sheet['C3'] = 3
    sheet['D3'] = 4
    sheet['A4'] = data.name + 1
    sheet['B4'] = data['Сотрудник']

    for column_index in range(4, 35):
        column_letter = get_column_letter(column_index)
        cell = sheet[f'{column_letter}{2}']
        cell.value = int(column_index - 3)

    # Merging
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    sheet.merge_cells('C1:C2')
    sheet.merge_cells('D1:AH1')
    sheet.merge_cells('D3:AH3')


def create_summary_sheet_header(workbook):
    sheet = workbook.worksheets[0]
    sheet.title = "Сводный лист"

    # Columns Width
    sheet.column_dimensions['A'].width = 6
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 15
    sheet.column_dimensions['D'].width = 12

    for i in range(5, 21):
        ltr = get_column_letter(i)
        sheet.column_dimensions[ltr].width = 3.4

    sheet.column_dimensions['U'].width = 8
    sheet.column_dimensions['V'].width = 8

    # Rows Height
    for i in range(1, 8):
        sheet.row_dimensions[i].height = 13

    for i in range(8, 13):
        sheet.row_dimensions[i].height = 16

    sheet.row_dimensions[13].height = 10

    # Style
    sheet['B1'].font = font2
    sheet['B1'].border = Border(bottom=Side(border_style='thin', color='000000'))
    sheet['B2'].font = font1
    sheet['B2'].alignment = text_style_manuscript
    sheet['B3'].font = font2
    sheet['B3'].border = Border(bottom=Side(border_style='thin', color='000000'))
    sheet['B4'].font = font1
    sheet['B4'].alignment = text_style_manuscript
    sheet['B7'].font = font3
    sheet['B7'].alignment = text_style

    sheet.freeze_panes = 'Y14'

    cell_range = sheet['B8:V13']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    # Data
    sheet['B1'] = 'Общество с ограниченной ответственностью "АСП-АКВА"'
    sheet['B2'] = 'наименование организации'
    sheet['B3'] = 'Название отдела Отдел трехмерного проектирования'
    sheet['B4'] = 'структурное поздразделение'
    sheet['B7'] = 'ТАБЕЛЬ'
    sheet['B8'] = 'Номер\n по\nпоряд-\nку'
    sheet['C8'] = 'Фамилия, инициалы,\n должность\n (специальность,\n профессия)'
    sheet['D8'] = 'Табельный\n номер'
    sheet['E8'] = 'Отметки о явках и неявках на работу по числам месяца'
    sheet['U8'] = 'Отработано за'

    for i in range(5, 20):
        ltr = get_column_letter(i)
        sheet[f'{ltr}9'] = i - 4
        sheet[f'{ltr}11'] = i + 11

    sheet['T9'] = sheet['T11'] = 'X'
    sheet['U9'] = 'половину\n месяца\n (I, II)'
    sheet['V9'] = 'месяц'
    sheet['U11'] = 'дни'
    sheet['U12'] = 'часы'
    sheet['B13'] = 1
    sheet['C13'] = 2
    sheet['D13'] = 3
    sheet['E13'] = 4
    sheet['U13'] = 5
    sheet['V13'] = 6

    # Merging
    sheet.merge_cells('B1:V1')
    sheet.merge_cells('B2:V2')
    sheet.merge_cells('B3:V3')
    sheet.merge_cells('B4:V4')
    sheet.merge_cells('B7:V7')
    sheet.merge_cells('B8:B12')
    sheet.merge_cells('C8:C12')
    sheet.merge_cells('D8:D12')
    sheet.merge_cells('E8:T8')
    sheet.merge_cells('U8:V8')

    for i in range(5, 21):
        ltr = get_column_letter(i)
        sheet.merge_cells(f'{ltr}9:{ltr}10')
        sheet.merge_cells(f'{ltr}11:{ltr}12')

    sheet.merge_cells('U9:U10')
    sheet.merge_cells('V9:V10')
    sheet.merge_cells('U11:V11')
    sheet.merge_cells('U12:V12')
    sheet.merge_cells('E13:T13')


def create_summary_sheet_footer(sheet, max_index):
    current_row = max_index + 1

    # Rows Height
    sheet.row_dimensions[current_row].height = 36
    sheet.row_dimensions[current_row + 3].height = 36

    # Style
    sheet[f'B{current_row}'].font = font4
    sheet[f'B{current_row}'].alignment = text_style_human
    sheet[f'S{current_row}'].font = font4
    sheet[f'S{current_row}'].alignment = text_style_human
    sheet[f'D{current_row}'].border = Border(bottom=Side(border_style='thin', color='000000'))
    sheet[f'F{current_row}'].border = Border(bottom=Side(border_style='thin', color='000000'))
    sheet[f'J{current_row}'].border = Border(bottom=Side(border_style='thin', color='000000'))
    sheet[f'D{current_row + 1}'].font = font1
    sheet[f'D{current_row + 1}'].alignment = text_style_manuscript
    sheet[f'F{current_row + 1}'].font = font1
    sheet[f'F{current_row + 1}'].alignment = text_style_manuscript
    sheet[f'J{current_row + 1}'].font = font1
    sheet[f'J{current_row + 1}'].alignment = text_style_manuscript
    sheet[f'V{current_row + 1}'].font = font1
    sheet[f'V{current_row + 1}'].alignment = text_style_manuscript
    sheet[f'S{current_row + 3}'].font = font4
    sheet[f'S{current_row + 3}'].alignment = text_style_human
    sheet[f'V{current_row + 4}'].font = font1
    sheet[f'V{current_row + 4}'].alignment = text_style_manuscript

    # Data
    sheet[f'B{current_row}'] = 'Ответственное\nлицо'
    sheet[f'S{current_row}'] = 'Руководитель\nструктурного\nподразделения'
    sheet[f'D{current_row + 1}'] = 'должность'
    sheet[f'F{current_row + 1}'] = 'личная подпись'
    sheet[f'J{current_row + 1}'] = 'расшифровка подписи'
    sheet[f'V{current_row + 1}'] = 'должность'
    sheet[f'S{current_row + 3}'] = 'Работник кадровой\nслужбы'
    sheet[f'V{current_row + 4}'] = 'должность'

    # Merging
    sheet.merge_cells(f'B{current_row}:C{current_row}')
    sheet.merge_cells(f'F{current_row}:H{current_row}')
    sheet.merge_cells(f'J{current_row}:Q{current_row}')
    sheet.merge_cells(f'S{current_row}:U{current_row}')
    sheet.merge_cells(f'F{current_row + 1}:H{current_row + 1}')
    sheet.merge_cells(f'J{current_row + 1}:Q{current_row + 1}')
    sheet.merge_cells(f'S{current_row + 3}:U{current_row + 3}')


def add_table(pos, sheet):

    # Rows Height
    for i in range(pos, pos+4):
        sheet.row_dimensions[i].height = 10

    # Style
    cell_range = sheet[f'B{pos}:V{pos + 3}']
    for row in cell_range:
        for cell in row:
            cell.alignment = text_style
            cell.font = font1
            cell.border = border

    # Merging
    sheet.merge_cells(f'B{pos}:B{pos + 3}')
    sheet.merge_cells(f'C{pos}:C{pos + 3}')
    sheet.merge_cells(f'D{pos}:D{pos + 3}')
    sheet.merge_cells(f'V{pos}:V{pos + 1}')
    sheet.merge_cells(f'V{pos+2}:V{pos + 3}')


def clear_summary_table(sheet):
    max_row = sheet.max_row
    if max_row > 14:
        sheet.unmerge_cells(f'B{max_row - 4}:C{max_row - 4}')
        sheet.unmerge_cells(f'F{max_row - 4}:H{max_row - 4}')
        sheet.unmerge_cells(f'J{max_row - 4}:Q{max_row - 4}')
        sheet.unmerge_cells(f'S{max_row - 4}:U{max_row - 4}')
        sheet.unmerge_cells(f'F{max_row - 3}:H{max_row - 3}')
        sheet.unmerge_cells(f'J{max_row - 3}:Q{max_row - 3}')
        sheet.unmerge_cells(f'S{max_row - 1}:U{max_row - 1}')
        sheet.delete_rows(14, max_row)


# Filling
def fill_employees_page(workbook, data, date):
    letter = get_column_letter(date + 3)

    for i in range(len(data)):
        if data.iloc[i]['Сотрудник'] not in workbook.sheetnames:
            create_employee_sheet(workbook, data.iloc[i])

    today_employees_list = data['Сотрудник'].tolist()

    for employee in workbook.sheetnames[1:]:
        if employee.title() not in today_employees_list:
            sheet = workbook[employee.title()]
            sheet[f'{letter}4'].fill = red_fill
        else:
            sheet = workbook[employee.title()]
            sheet[f'{letter}4'] = data.loc[data['Сотрудник'] == employee.title()]['Время, часы'].item()


def fill_summary_table(workbook, old_data):
    sheet = workbook.worksheets[0]
    clear_summary_table(sheet)

    for employee_sheet in workbook.worksheets[1:]:
        index = employee_sheet['A4'].value
        name = employee_sheet['B4'].value
        pos = 10 + (4 * index)

        add_table(pos, sheet)

        sheet[f'B{pos}'] = index
        sheet[f'C{pos}'] = name

        restore_old_data(name, sheet, index, old_data)

        for i in range(1, 31):
            letter_for_employee = get_column_letter(i+3)
            time = employee_sheet[f'{letter_for_employee}4'].value
            filled = employee_sheet[f'{letter_for_employee}4'].fill
            if time is not None and type(time) is not str:
                if i < 16:
                    sheet[f'{get_column_letter(i+4)}{pos}'] = 'Я'
                    sheet[f'{get_column_letter(i+4)}{pos + 1}'] = '' if time == 8 else time - 8
                else:
                    sheet[f'{get_column_letter(i-11)}{pos + 2}'] = 'Я'
                    sheet[f'{get_column_letter(i-11)}{pos + 3}'] = '' if time == 8 else time - 8
            elif filled is not None:
                fill = PatternFill(fill_type=filled.fill_type, fgColor=filled.fgColor, bgColor=filled.bgColor)
                if i < 16:
                    sheet[f'{get_column_letter(i + 4)}{pos}'] = time
                    sheet[f'{get_column_letter(i + 4)}{pos}'].fill = fill
                    sheet[f'{get_column_letter(i + 4)}{pos + 1}'].fill = fill
                else:
                    sheet[f'{get_column_letter(i - 11)}{pos + 2}'] = time
                    sheet[f'{get_column_letter(i - 11)}{pos + 2}'].fill = fill
                    sheet[f'{get_column_letter(i - 11)}{pos + 3}'].fill = fill

    max_row = sheet.max_row
    create_summary_sheet_footer(sheet, max_row)


# To Remember and to Restore cells
def to_remember(workbook):
    sheet = workbook.worksheets[0]
    employees = len(workbook.worksheets) - 1
    data = {}
    if employees > 0:
        for employee_sheet in workbook.worksheets[1:]:
            index = employee_sheet['A4'].value
            name = employee_sheet['B4'].value
            pos = 10 + (4 * index)
            cell_range = sheet[f'E{pos}:T{pos + 3}']
            employee_data = {'index': index,
                             'coordinates' : {},
                             'design': {}}
            for row in cell_range:
                for cell in row:
                    value = cell.value
                    if value is not None:
                        employee_data['coordinates'][cell.coordinate] = value
                        employee_data['design'][cell.coordinate] = cell.fill
            data[name] = employee_data
    return data


def restore_old_data(name, sheet, index, old_data):
    if len(old_data) > 0 and name in old_data.keys():
        coordinates = old_data[name]['coordinates']
        for coordinate in coordinates:
            column = coordinate[:1]
            row = int(coordinate[1:])
            index_dif = index - int(old_data[name]['index'])
            rows_move = 4 * index_dif
            row += rows_move
            new_coordinate = column + str(row)
            sheet[new_coordinate] = old_data[name]['coordinates'][coordinate]

            filled = old_data[name]['design'][coordinate]
            fill = PatternFill(fill_type=filled.fill_type, fgColor=filled.fgColor, bgColor=filled.bgColor)
            sheet[new_coordinate].fill = fill


# Building Reports
def create_report(excel_file, selected_date):
    workbook = load_workbook(excel_file)

    employees = []
    for employee_page in workbook.worksheets[1:]:
        employees.append(employee_page.title)
    data = create_dataframe(selected_date, employees)
    day = int(str(selected_date)[-2:])
    df = count_hours_per_day(data)
    old_data = to_remember(workbook)
    fill_employees_page(workbook, df, day)

    create_summary_sheet_header(workbook)
    re_index(workbook)
    sort_pages(workbook)
    fill_summary_table(workbook, old_data)

    workbook.save(excel_file)
    print('Готово!')


def prebuild(report, staff):
    workbook = load_workbook(report)
    old_data = to_remember(workbook)

    df = pd.read_excel(staff, usecols=['ФИО',])
    df.rename(columns={'ФИО': 'Сотрудник'}, inplace=True)
    df = df.dropna()

    for i in range(len(df)):
        df.loc[i]['Сотрудник'] = shorten_name(df.loc[i]['Сотрудник'])
        if df.loc[i]['Сотрудник'] not in workbook.sheetnames:
            create_employee_sheet(workbook, df.loc[i])

    for employee_sheet in workbook.sheetnames[1:]:
        if employee_sheet.title() not in list(df['Сотрудник']):
            std = workbook.get_sheet_by_name(employee_sheet.title())
            workbook.remove_sheet(std)

    if 'Сводный лист' in workbook.sheetnames:
        del workbook['Сводный лист']

    workbook.create_sheet('Сводный лист', 0)

    if 'Лист1' in workbook.sheetnames:
        del workbook['Сводный лист']

    create_summary_sheet_header(workbook)
    re_index(workbook)
    sort_pages(workbook)
    fill_summary_table(workbook, old_data)

    workbook.save(report)
    print('Пре-Билд готов!')


def reload_report(excel_file):
    workbook = load_workbook(excel_file)
    old_data = to_remember(workbook)
    fill_summary_table(workbook, old_data)
    workbook.save(excel_file)


# Other
def count_hours_per_day(data):
    table = data.groupby('Сотрудник')['Время, часы'].sum().reset_index()
    table = table.sort_values(by='Сотрудник', key=lambda x: x.str.split().str[-1])
    result = table.reset_index(drop=True)
    return result


def re_index(workbook):
    sheets = workbook.worksheets
    names = []
    if len(sheets) > 1:
        for sheet in sheets[1:]:
            names.append(sheet['B4'].value)
        names.sort(key=lambda name: name.split()[-1])
        for sheet in sheets[1:]:
            sheet['A4'].value = names.index(sheet.title) + 1


def sort_pages(workbook):
    sorted_sheets = sorted(workbook._sheets[1:], key=lambda sheet: sheet.title.split()[-1])
    workbook._sheets[1:] = sorted_sheets


def shorten_name(full_name):
    split_names = full_name.split()
    shortened_name = split_names[1] + ' ' + split_names[0]
    return shortened_name
